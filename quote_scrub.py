import openpyxl
import os
import sys


def scrub(quote_file):

    # check for file
    if not os.path.isfile(quote_file):
        print("Could not find quote file " + quote_file, file=sys.stderr)
        return

    # read quote from quote.xlsx
    wb = openpyxl.load_workbook(quote_file, read_only=True)
    sheet = wb.active
    rows = sheet.rows

    # column headings in header_row
    header_row = [cell.value for cell in next(rows)]

    quote = []

    # put remaining rows in quote
    for row in rows:
        record = {}

        # store each row in record
        for key, cell in zip(header_row, row):

            if key == "Serial #":
                # save serial #'s as strings, strip off spaces
                record[key] = str(cell.value).strip()

            elif cell.data_type == 's':
                # strip extra spaces from strings
                record[key] = cell.value.strip()

            else:
                # store everything else (numbers and dates) unchanged
                record[key] = cell.value

        # add row/record to quote
        quote.append(record)

    # prepare to write quote_scrubed.xlsx
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    # add column headings
    ws.append(header_row)

    # keep running total of Net Price
    net_price_total = 0
    net_price_by_serial = {}
    for row in quote:

        if ',' in row["Serial #"]:
            # multiple serial numbers in this row, need to split
            serials = row["Serial #"].split(',')

            for serial in serials:
                # print(serial.strip())
                new_row = dict(row)
                new_row["Serial #"] = serial.strip()
                new_row["Qty"] = 1
                new_row["Ext Qty"] = 1
                new_row["Ext List Price"] = row["List Price"]
                new_row["Ext Net Price"] = row["Net Price"]
                ws.append([new_row[column] for column in header_row])
                net_price_total += row["Net Price"]
                if serial.strip() not in net_price_by_serial:
                    net_price_by_serial[serial.strip()] = 0
                net_price_by_serial[serial.strip()] += row["Net Price"]
        else:
            # single serial number in this row
            ws.append([row[column] for column in header_row])
            net_price_total += row["Net Price"]
            serial = row["Serial #"].strip()
            if serial not in net_price_by_serial:
                net_price_by_serial[serial] = 0
            net_price_by_serial[serial] += row["Net Price"]

    # insert total net price
    ws.append([])
    ws.append(["Total Net Price:"])
    ws.append([net_price_total])

    ws = wb.create_sheet()
    ws.append(["Serial #", "Net Price"])
    for serial in net_price_by_serial:
        print(serial + " total net price: " + str(net_price_by_serial[serial]))
        ws.append([serial, net_price_by_serial[serial]])

    save_file_name = quote_file.replace('.xlsx', '_scrubed.xlsx')
    wb.save(save_file_name)
    print("Done, created " + save_file_name, file=sys.stderr)
if __name__ == "__main__":
    print("Attempting to scrub quote.xlsx", file=sys.stderr)
    scrub("quote.xlsx")
